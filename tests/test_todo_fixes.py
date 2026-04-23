from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.core.database import get_engine
from app.models.ai_audit_report import AIAuditReport
from app.models.refresh_token import RefreshToken

API_PREFIX = "/api/v1"


def assert_ok(response):
    payload = response.json()
    assert response.status_code == 200, payload
    assert payload["code"] == 0, payload
    return payload["data"]


def auth_headers(access_token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {access_token}"}


def register_user(
    client,
    *,
    account: str,
    password: str = "pass1234",
    name: str = "Test User",
    role: str = "student",
    class_id: int | None = None,
    email: str | None = None,
):
    response = client.post(
        f"{API_PREFIX}/auth/register",
        json={
            "account": account,
            "password": password,
            "name": name,
            "role": role,
            "class_id": class_id,
            "email": email,
        },
    )
    return assert_ok(response)["user"]


def login_user(client, *, account: str, password: str = "pass1234"):
    response = client.post(
        f"{API_PREFIX}/auth/login",
        json={"account": account, "password": password},
    )
    return assert_ok(response)


def upload_file(client, access_token: str, filename: str = "proof.png", content_type: str = "image/png") -> str:
    body = b"\x89PNG\r\n\x1a\nmock" if filename.endswith(".png") else b"mock"
    response = client.post(
        f"{API_PREFIX}/files/upload",
        headers=auth_headers(access_token),
        files={"file": (filename, body, content_type)},
    )
    return assert_ok(response)["file_id"]


def create_application(client, access_token: str, *, title: str, file_id: str, score: float = 4.0):
    response = client.post(
        f"{API_PREFIX}/applications",
        headers=auth_headers(access_token),
        json={
            "award_uid": 1,
            "title": title,
            "description": f"description {title}",
            "occurred_at": date.today().isoformat(),
            "attachments": [{"file_id": file_id}],
            "category": "innovation",
            "sub_type": "achievement",
            "score": score,
        },
    )
    return assert_ok(response)


def create_application_custom(
    client,
    access_token: str,
    *,
    title: str,
    file_id: str,
    category: str,
    sub_type: str,
    score: float = 4.0,
):
    response = client.post(
        f"{API_PREFIX}/applications",
        headers=auth_headers(access_token),
        json={
            "award_uid": 1,
            "title": title,
            "description": f"description {title}",
            "occurred_at": date.today().isoformat(),
            "attachments": [{"file_id": file_id}],
            "category": category,
            "sub_type": sub_type,
            "score": score,
        },
    )
    return assert_ok(response)


def test_users_me_returns_tokens(client):
    register_user(client, account="todo_teacher", role="teacher", name="Todo Teacher")
    register_user(client, account="todo_student", class_id=301, name="Todo Student")
    register_user(client, account="todo_reviewer", class_id=301, name="Todo Reviewer")

    teacher_login = login_user(client, account="todo_teacher")
    teacher_headers = auth_headers(teacher_login["access_token"])

    created_token = assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer",
            headers=teacher_headers,
            json={"class_ids": [301]},
        )
    )

    teacher_me = assert_ok(client.get(f"{API_PREFIX}/users/me", headers=teacher_headers))
    assert isinstance(teacher_me["tokens"], list)
    assert len(teacher_me["tokens"]) == 1
    assert teacher_me["tokens"][0]["token"] == created_token["token"]

    reviewer_login = login_user(client, account="todo_reviewer")
    reviewer_headers = auth_headers(reviewer_login["access_token"])
    assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer/activate",
            headers=reviewer_headers,
            json={"token": created_token["token"]},
        )
    )

    reviewer_me = assert_ok(client.get(f"{API_PREFIX}/users/me", headers=reviewer_headers))
    assert len(reviewer_me["tokens"]) == 1
    assert reviewer_me["tokens"][0]["status"] == "active"
    assert reviewer_me["tokens"][0]["activated_user_id"] == reviewer_me["id"]

    student_login = login_user(client, account="todo_student")
    student_me = assert_ok(client.get(f"{API_PREFIX}/users/me", headers=auth_headers(student_login["access_token"])))
    assert student_me["tokens"] == []


def test_teacher_student_statistics_and_export_summary_archive(client):
    register_user(client, account="todo_teacher2", role="teacher", name="Teacher 2")
    register_user(client, account="todo_reviewer2", class_id=301, name="Reviewer 2")
    register_user(client, account="todo_student_a", class_id=301, name="Student A")
    register_user(client, account="todo_student_b", class_id=301, name="Student B")

    teacher_login = login_user(client, account="todo_teacher2")
    teacher_headers = auth_headers(teacher_login["access_token"])

    token_data = assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer",
            headers=teacher_headers,
            json={"class_ids": [301]},
        )
    )

    reviewer_login = login_user(client, account="todo_reviewer2")
    reviewer_headers = auth_headers(reviewer_login["access_token"])
    assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer/activate",
            headers=reviewer_headers,
            json={"token": token_data["token"]},
        )
    )

    student_a_login = login_user(client, account="todo_student_a")
    student_b_login = login_user(client, account="todo_student_b")
    file_a = upload_file(client, student_a_login["access_token"])
    file_b = upload_file(client, student_b_login["access_token"])
    app_a = create_application(client, student_a_login["access_token"], title="A", file_id=file_a, score=4.5)
    app_b = create_application(client, student_b_login["access_token"], title="B", file_id=file_b, score=3.0)

    assert_ok(
        client.post(
            f"{API_PREFIX}/reviews/{app_a['application_id']}/decision",
            headers=reviewer_headers,
            json={"decision": "approved", "comment": "ok"},
        )
    )
    assert_ok(
        client.post(
            f"{API_PREFIX}/reviews/{app_b['application_id']}/decision",
            headers=reviewer_headers,
            json={"decision": "approved", "comment": "ok"},
        )
    )

    assert_ok(
        client.post(
            f"{API_PREFIX}/teacher/applications/{app_a['application_id']}/recheck",
            headers=teacher_headers,
            json={"decision": "approved", "comment": "final ok", "score": 4.5},
        )
    )
    assert_ok(
        client.post(
            f"{API_PREFIX}/teacher/applications/{app_b['application_id']}/recheck",
            headers=teacher_headers,
            json={"decision": "rejected", "comment": "final reject", "score": 3.0},
        )
    )
    assert_ok(
        client.post(
            f"{API_PREFIX}/teacher/applications/archive",
            headers=teacher_headers,
            json={"application_ids": [app_a["application_id"], app_b["application_id"]]},
        )
    )

    student_stats = assert_ok(
        client.get(
            f"{API_PREFIX}/teacher/statistics/students?grade=2023&class_id=301",
            headers=teacher_headers,
        )
    )
    assert len(student_stats["list"]) == 2
    stats_by_account = {item["student_account"]: item for item in student_stats["list"]}
    assert stats_by_account["todo_student_a"]["actual_score"] == 4.5
    assert stats_by_account["todo_student_b"]["actual_score"] == 0.0
    for row in student_stats["list"]:
        assert set(row).issuperset(
            {
                "grade",
                "class_id",
                "student_id",
                "student_account",
                "student_name",
                "total_count",
                "rejected_count",
                "pending_count",
                "total_score",
                "average_score",
                "actual_score",
            }
        )

    export_task = assert_ok(
        client.post(
            f"{API_PREFIX}/teacher/exports",
            headers={**teacher_headers, "Idempotency-Key": "todo-export-students-001"},
            json={
                "scope": "teacher_statistics",
                "format": "xlsx",
                "filters": {"grade": 2023, "class_id": 301, "term": "2025-2026-1"},
                "store_to_archive": True,
            },
        )
    )
    task_id = export_task["task_id"]

    export_detail = assert_ok(client.get(f"{API_PREFIX}/teacher/exports/{task_id}", headers=teacher_headers))
    assert export_detail["status"] == "completed"
    assert export_detail["file_name"] == "teacher_statistics_2025-2026-1.xlsx"
    assert export_detail["file_url"]

    download = client.get(export_detail["file_url"], headers=teacher_headers)
    assert download.status_code == 200

    workbook = load_workbook(BytesIO(download.content))
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers == [
        "grade",
        "class_id",
        "student_id",
        "student_account",
        "student_name",
        "total_count",
        "rejected_count",
        "pending_count",
        "total_score",
        "average_score",
        "actual_score",
    ]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2

    archives = assert_ok(client.get(f"{API_PREFIX}/archives/exports", headers=teacher_headers))
    matched = [item for item in archives if item["export_task_id"] == task_id]
    assert matched


def test_upload_docx_supported(client):
    register_user(client, account="todo_docx_student", class_id=301, name="Docx Student")
    login_data = login_user(client, account="todo_docx_student")
    response = client.post(
        f"{API_PREFIX}/files/upload",
        headers=auth_headers(login_data["access_token"]),
        files={
            "file": (
                "proof.docx",
                b"PK\x03\x04mock-docx-content",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        },
    )
    data = assert_ok(response)
    assert data["file_id"].endswith(".docx")


def test_reviewer_multi_tokens_union_and_unbind(client):
    register_user(client, account="todo_teacher_multi", role="teacher", name="Teacher Multi")
    register_user(client, account="todo_reviewer_multi", class_id=301, name="Reviewer Multi")
    register_user(client, account="todo_student_cls301", class_id=301, name="Student 301")
    register_user(client, account="todo_student_cls302", class_id=302, name="Student 302")

    teacher_login = login_user(client, account="todo_teacher_multi")
    teacher_headers = auth_headers(teacher_login["access_token"])
    reviewer_login = login_user(client, account="todo_reviewer_multi")
    reviewer_headers = auth_headers(reviewer_login["access_token"])

    token_301 = assert_ok(
        client.post(f"{API_PREFIX}/tokens/reviewer", headers=teacher_headers, json={"class_ids": [301]})
    )
    token_302 = assert_ok(
        client.post(f"{API_PREFIX}/tokens/reviewer", headers=teacher_headers, json={"class_ids": [302]})
    )

    assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer/activate",
            headers=reviewer_headers,
            json={"token": token_301["token"]},
        )
    )
    assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer/activate",
            headers=reviewer_headers,
            json={"token": token_302["token"]},
        )
    )

    reviewer_tokens = assert_ok(client.get(f"{API_PREFIX}/tokens?type=reviewer", headers=reviewer_headers))
    assert reviewer_tokens["total"] == 2

    student_301_login = login_user(client, account="todo_student_cls301")
    student_302_login = login_user(client, account="todo_student_cls302")
    app_301 = create_application(
        client,
        student_301_login["access_token"],
        title="multi-301",
        file_id=upload_file(client, student_301_login["access_token"]),
        score=4.0,
    )
    app_302 = create_application(
        client,
        student_302_login["access_token"],
        title="multi-302",
        file_id=upload_file(client, student_302_login["access_token"]),
        score=5.0,
    )

    pending_all = assert_ok(client.get(f"{API_PREFIX}/reviews/pending", headers=reviewer_headers))
    ids = {row["application_id"] for row in pending_all["list"]}
    assert app_301["application_id"] in ids
    assert app_302["application_id"] in ids

    assert_ok(client.post(f"{API_PREFIX}/tokens/{token_302['token_id']}/unbind", headers=reviewer_headers))
    pending_after_unbind = assert_ok(client.get(f"{API_PREFIX}/reviews/pending", headers=reviewer_headers))
    ids_after_unbind = {row["application_id"] for row in pending_after_unbind["list"]}
    assert app_301["application_id"] in ids_after_unbind
    assert app_302["application_id"] not in ids_after_unbind


def test_reviewer_can_view_ai_report_in_bound_class_scope(client):
    register_user(client, account="todo_teacher_ai_scope", role="teacher", name="Teacher AI Scope")
    register_user(client, account="todo_reviewer_ai_scope", class_id=301, name="Reviewer AI Scope")
    register_user(client, account="todo_student_ai_301", class_id=301, name="Student AI 301")
    register_user(client, account="todo_student_ai_302", class_id=302, name="Student AI 302")

    teacher_login = login_user(client, account="todo_teacher_ai_scope")
    teacher_headers = auth_headers(teacher_login["access_token"])
    reviewer_login = login_user(client, account="todo_reviewer_ai_scope")
    reviewer_headers = auth_headers(reviewer_login["access_token"])

    token = assert_ok(client.post(f"{API_PREFIX}/tokens/reviewer", headers=teacher_headers, json={"class_ids": [301]}))
    assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer/activate",
            headers=reviewer_headers,
            json={"token": token["token"]},
        )
    )

    student_301_login = login_user(client, account="todo_student_ai_301")
    student_302_login = login_user(client, account="todo_student_ai_302")

    app_301 = create_application(
        client,
        student_301_login["access_token"],
        title="ai-scope-301",
        file_id=upload_file(client, student_301_login["access_token"]),
        score=4.0,
    )
    app_302 = create_application(
        client,
        student_302_login["access_token"],
        title="ai-scope-302",
        file_id=upload_file(client, student_302_login["access_token"]),
        score=4.0,
    )

    # ensure report rows exist before reviewer query
    with Session(get_engine()) as db:
        assert db.exec(select(AIAuditReport).where(AIAuditReport.application_id == app_301["application_id"])).first()
        assert db.exec(select(AIAuditReport).where(AIAuditReport.application_id == app_302["application_id"])).first()

    report_in_scope = assert_ok(
        client.get(f"{API_PREFIX}/ai-audits/{app_301['application_id']}/report", headers=reviewer_headers)
    )
    assert report_in_scope["application_id"] == app_301["application_id"]

    out_scope_resp = client.get(f"{API_PREFIX}/ai-audits/{app_302['application_id']}/report", headers=reviewer_headers)
    out_scope_payload = out_scope_resp.json()
    assert out_scope_resp.status_code == 200
    assert out_scope_payload["code"] == 1003


def test_pending_count_scope_and_announcement_visibility(client):
    register_user(client, account="todo_teacher_scope", role="teacher", name="Teacher Scope")
    register_user(client, account="todo_reviewer_scope", class_id=301, name="Reviewer Scope")
    register_user(client, account="todo_student_scope", class_id=301, name="Student Scope")

    teacher_login = login_user(client, account="todo_teacher_scope")
    teacher_headers = auth_headers(teacher_login["access_token"])
    reviewer_login = login_user(client, account="todo_reviewer_scope")
    reviewer_headers = auth_headers(reviewer_login["access_token"])
    student_login = login_user(client, account="todo_student_scope")
    student_headers = auth_headers(student_login["access_token"])

    token = assert_ok(client.post(f"{API_PREFIX}/tokens/reviewer", headers=teacher_headers, json={"class_ids": [301]}))
    assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer/activate",
            headers=reviewer_headers,
            json={"token": token["token"]},
        )
    )

    file_a = upload_file(client, student_login["access_token"], filename="a.png", content_type="image/png")
    file_b = upload_file(client, student_login["access_token"], filename="b.png", content_type="image/png")
    create_application_custom(
        client,
        student_login["access_token"],
        title="innovation-achievement",
        file_id=file_a,
        category="innovation",
        sub_type="achievement",
    )
    create_application_custom(
        client,
        student_login["access_token"],
        title="art-basic",
        file_id=file_b,
        category="art",
        sub_type="basic",
    )

    pending_all = assert_ok(client.get(f"{API_PREFIX}/reviews/pending-count", headers=reviewer_headers))
    assert pending_all["pending_count"] == 2

    pending_innovation = assert_ok(
        client.get(
            f"{API_PREFIX}/reviews/pending-count?category=innovation&sub_type=achievement",
            headers=reviewer_headers,
        )
    )
    assert pending_innovation["pending_count"] == 1

    export_task = assert_ok(
        client.post(
            f"{API_PREFIX}/archives/exports",
            headers=teacher_headers,
            json={
                "scope": "public_archive",
                "format": "xlsx",
                "filters": {"term": "2025-2026-1", "grade": 2023, "class_id": 301},
                "store_to_archive": True,
            },
        )
    )
    task_id = export_task["task_id"]
    archives = assert_ok(client.get(f"{API_PREFIX}/archives/exports", headers=teacher_headers))
    archive = next(item for item in archives if item["export_task_id"] == task_id)

    announcement = assert_ok(
        client.post(
            f"{API_PREFIX}/announcements",
            headers=teacher_headers,
            json={
                "title": "scope announcement",
                "archive_id": archive["archive_id"],
                "scope": {"grade": 2023, "class_ids": [301]},
                "start_at": "2026-04-01T00:00:00+08:00",
                "end_at": "2026-05-01T00:00:00+08:00",
            },
        )
    )
    announcement_id = announcement["announcement_id"]

    visible_before_close = assert_ok(client.get(f"{API_PREFIX}/announcements", headers=student_headers))
    assert any(item["announcement_id"] == announcement_id for item in visible_before_close)

    assert_ok(client.post(f"{API_PREFIX}/announcements/{announcement_id}/close", headers=teacher_headers))

    visible_after_close = assert_ok(client.get(f"{API_PREFIX}/announcements", headers=student_headers))
    assert all(item["announcement_id"] != announcement_id for item in visible_after_close)


def test_total_score_excludes_rejected_status(client):
    register_user(client, account="todo_teacher_score", role="teacher", name="Teacher Score")
    register_user(client, account="todo_reviewer_score", class_id=301, name="Reviewer Score")
    register_user(client, account="todo_student_score_a", class_id=301, name="Student Score A")
    register_user(client, account="todo_student_score_b", class_id=301, name="Student Score B")

    teacher_login = login_user(client, account="todo_teacher_score")
    teacher_headers = auth_headers(teacher_login["access_token"])
    reviewer_login = login_user(client, account="todo_reviewer_score")
    reviewer_headers = auth_headers(reviewer_login["access_token"])

    token = assert_ok(client.post(f"{API_PREFIX}/tokens/reviewer", headers=teacher_headers, json={"class_ids": [301]}))
    assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer/activate",
            headers=reviewer_headers,
            json={"token": token["token"]},
        )
    )

    student_a_login = login_user(client, account="todo_student_score_a")
    student_b_login = login_user(client, account="todo_student_score_b")
    app_a = create_application(
        client,
        student_a_login["access_token"],
        title="score-a",
        file_id=upload_file(client, student_a_login["access_token"]),
        score=4.0,
    )
    app_b = create_application(
        client,
        student_b_login["access_token"],
        title="score-b",
        file_id=upload_file(client, student_b_login["access_token"]),
        score=3.0,
    )

    assert_ok(
        client.post(
            f"{API_PREFIX}/reviews/{app_a['application_id']}/decision",
            headers=reviewer_headers,
            json={"decision": "approved", "comment": "ok"},
        )
    )
    assert_ok(
        client.post(
            f"{API_PREFIX}/reviews/{app_b['application_id']}/decision",
            headers=reviewer_headers,
            json={"decision": "rejected", "comment": "reject"},
        )
    )

    student_stats = assert_ok(
        client.get(f"{API_PREFIX}/teacher/statistics/students?grade=2023&class_id=301", headers=teacher_headers)
    )
    stats_by_account = {item["student_account"]: item for item in student_stats["list"]}
    assert stats_by_account["todo_student_score_a"]["total_score"] == 4.0
    assert stats_by_account["todo_student_score_b"]["total_score"] == 0.0


def test_update_profile_accepts_blank_email_and_phone(client):
    register_user(
        client,
        account="todo_profile_blank",
        class_id=301,
        name="Profile Blank",
        email="profile.blank@example.com",
    )
    login_data = login_user(client, account="todo_profile_blank")
    headers = auth_headers(login_data["access_token"])

    payload = assert_ok(client.put(f"{API_PREFIX}/users/me", headers=headers, json={"email": "", "phone": ""}))
    assert payload["email"] is None
    assert payload["phone"] is None


def test_update_profile_accepts_legacy_local_email(client):
    register_user(
        client,
        account="todo_profile_local",
        class_id=301,
        name="Profile Local",
        email="profile.local@example.com",
    )
    login_data = login_user(client, account="todo_profile_local")
    headers = auth_headers(login_data["access_token"])

    payload = assert_ok(
        client.put(
            f"{API_PREFIX}/users/me",
            headers=headers,
            json={"email": "profile.local@zcpt.local", "phone": "18800000000"},
        )
    )
    assert payload["email"] == "profile.local@zcpt.local"
    assert payload["phone"] == "18800000000"


def test_activate_expired_token_returns_business_error(client):
    register_user(client, account="todo_teacher_expired", role="teacher", name="Teacher Expired")
    register_user(client, account="todo_student_expired", class_id=301, name="Student Expired")

    teacher_login = login_user(client, account="todo_teacher_expired")
    teacher_headers = auth_headers(teacher_login["access_token"])
    student_login = login_user(client, account="todo_student_expired")
    student_headers = auth_headers(student_login["access_token"])

    expired_at = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
    token = assert_ok(
        client.post(
            f"{API_PREFIX}/tokens/reviewer",
            headers=teacher_headers,
            json={"class_ids": [301], "expired_at": expired_at},
        )
    )

    response = client.post(
        f"{API_PREFIX}/tokens/reviewer/activate",
        headers=student_headers,
        json={"token": token["token"]},
    )
    payload = response.json()
    assert response.status_code == 200
    assert payload["code"] == 1000
    assert payload["message"] == "token expired"


def test_auth_refresh_handles_naive_expires_at(client):
    register_user(client, account="todo_refresh_naive", class_id=301, name="Refresh Naive")
    login_data = login_user(client, account="todo_refresh_naive")

    refresh_token = login_data["refresh_token"]

    with Session(get_engine()) as db:
        row = db.exec(select(RefreshToken).order_by(RefreshToken.id.desc())).first()
        assert row is not None
        # simulate legacy sqlite row that stores naive datetime
        row.expires_at = (datetime.now(timezone.utc) + timedelta(days=1)).replace(tzinfo=None)
        db.add(row)
        db.commit()

    refresh_resp = client.post(f"{API_PREFIX}/auth/refresh", json={"refresh_token": refresh_token})
    refreshed = assert_ok(refresh_resp)
    assert isinstance(refreshed.get("access_token"), str)
    assert refreshed.get("expires_in")


def test_announcement_status_binding_and_reopen(client):
    register_user(client, account="todo_teacher_announcement", role="teacher", name="Teacher Announcement")
    register_user(client, account="todo_student_announcement", class_id=301, name="Student Announcement")

    teacher_login = login_user(client, account="todo_teacher_announcement")
    teacher_headers = auth_headers(teacher_login["access_token"])
    student_login = login_user(client, account="todo_student_announcement")
    student_headers = auth_headers(student_login["access_token"])

    export_task = assert_ok(
        client.post(
            f"{API_PREFIX}/archives/exports",
            headers=teacher_headers,
            json={
                "scope": "public_archive",
                "format": "xlsx",
                "filters": {"term": "2025-2026-1", "grade": 2023, "class_id": 301},
                "store_to_archive": True,
            },
        )
    )
    archives = assert_ok(client.get(f"{API_PREFIX}/archives/exports", headers=teacher_headers))
    archive = next(item for item in archives if item["export_task_id"] == export_task["task_id"])

    now = datetime.now(timezone.utc)
    active_start = (now - timedelta(days=10)).isoformat()
    active_end = (now - timedelta(days=5)).isoformat()
    future_start = (now + timedelta(days=2)).isoformat()
    future_end = (now + timedelta(days=3)).isoformat()

    active_announcement = assert_ok(
        client.post(
            f"{API_PREFIX}/announcements",
            headers=teacher_headers,
            json={
                "title": "active announcement",
                "archive_id": archive["archive_id"],
                "scope": {"grade": 2023, "class_ids": [301]},
                "start_at": active_start,
                "end_at": active_end,
            },
        )
    )
    future_window_announcement = assert_ok(
        client.post(
            f"{API_PREFIX}/announcements",
            headers=teacher_headers,
            json={
                "title": "future window announcement",
                "archive_id": archive["archive_id"],
                "scope": {"grade": 2023, "class_ids": [301]},
                "start_at": future_start,
                "end_at": future_end,
            },
        )
    )

    visible = assert_ok(client.get(f"{API_PREFIX}/announcements", headers=student_headers))
    visible_ids = {item["announcement_id"] for item in visible}
    assert active_announcement["announcement_id"] in visible_ids
    assert future_window_announcement["announcement_id"] in visible_ids

    # active announcement can be appealed while active
    created_appeal = assert_ok(
        client.post(
            f"{API_PREFIX}/appeals",
            headers=student_headers,
            json={
                "announcement_id": active_announcement["announcement_id"],
                "content": "appeal before close",
                "attachments": [],
            },
        )
    )
    assert created_appeal["announcement_id"] == active_announcement["announcement_id"]

    assert_ok(client.post(f"{API_PREFIX}/announcements/{active_announcement['announcement_id']}/close", headers=teacher_headers))
    hidden_after_close = assert_ok(client.get(f"{API_PREFIX}/announcements", headers=student_headers))
    hidden_ids = {item["announcement_id"] for item in hidden_after_close}
    assert active_announcement["announcement_id"] not in hidden_ids

    denied_appeal_resp = client.post(
        f"{API_PREFIX}/appeals",
        headers=student_headers,
        json={
            "announcement_id": active_announcement["announcement_id"],
            "content": "appeal after close",
            "attachments": [],
        },
    )
    denied_payload = denied_appeal_resp.json()
    assert denied_appeal_resp.status_code == 200
    assert denied_payload["code"] == 1003

    assert_ok(
        client.post(f"{API_PREFIX}/announcements/{active_announcement['announcement_id']}/reopen", headers=teacher_headers)
    )
    visible_after_reopen = assert_ok(client.get(f"{API_PREFIX}/announcements", headers=student_headers))
    visible_after_reopen_ids = {item["announcement_id"] for item in visible_after_reopen}
    assert active_announcement["announcement_id"] in visible_after_reopen_ids
