from openpyxl.utils import get_column_letter

from app.core.score_rules import SCORE_CATEGORY_KEYS


EXPORT_SCORE_COLUMNS = [
    ("身心素养总分", "physical_mental_score"),
    ("身心素养基础分", "physical_mental_basic"),
    ("身心素养成果分", "physical_mental_achievement"),
    ("身心素养成果分+溢出分", "physical_mental_achievement_with_overflow"),
    ("文艺素养总分", "art_score"),
    ("文艺素养基础分", "art_basic"),
    ("文艺素养成果分", "art_achievement"),
    ("文艺素养成果分+溢出分", "art_achievement_with_overflow"),
    ("劳动素养总分", "labor_score"),
    ("劳动素养基础分", "labor_basic"),
    ("劳动素养成果分", "labor_achievement"),
    ("劳动素养成果分+溢出分", "labor_achievement_with_overflow"),
    ("创新素养总分", "innovation_score"),
    ("创新素养基础分", "innovation_basic"),
    ("创新素养突破分", "innovation_achievement"),
    ("创新素养突破分+溢出分", "innovation_achievement_with_overflow"),
]


def build_score_export_columns(score_summary: dict) -> dict[str, float]:
    category_scores = score_summary.get("category_scores") or {}
    sub_scores = score_summary.get("sub_scores") or {}
    achievement_overflow_scores = score_summary.get("achievement_overflow_scores") or {}
    values: dict[str, float] = {}
    for category in SCORE_CATEGORY_KEYS:
        values[f"{category}_score"] = _score(category_scores.get(f"{category}_score"))
        values[f"{category}_basic"] = _score(sub_scores.get(f"{category}_basic"))
        values[f"{category}_achievement"] = _score(sub_scores.get(f"{category}_achievement"))
        achievement_overflow = _score(achievement_overflow_scores.get(f"{category}_achievement_overflow"))
        values[f"{category}_achievement_with_overflow"] = round(
            values[f"{category}_achievement"] + achievement_overflow,
            4,
        )
    return values


def autosize_workbook_columns(workbook, *, min_width: int = 10, max_width: int = 34) -> None:
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row and worksheet.max_column:
            worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                max_length = max(max_length, _display_width(cell.value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def _display_width(value) -> int:
    text = "" if value is None else str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def _score(value) -> float:
    return round(float(value or 0.0), 4)
