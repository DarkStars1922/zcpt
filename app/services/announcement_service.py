from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.exc import IntegrityError
from sqlmodel import Session, select

from app.core.award_catalog import serialize_award_rule
from app.core.constants import ANNOUNCEMENT_STATUS_ACTIVE, ANNOUNCEMENT_STATUS_CLOSED
from app.core.score_rules import SCORE_CATEGORY_KEYS, SCORE_CATEGORY_RULES, SCORE_SUB_TYPE_KEYS
from app.core.config import settings
from app.core.term_utils import datetime_in_term, format_term_label
from app.core.utils import ensure_utc_datetime, json_dumps, json_loads, utcnow
from app.models.announcement import Announcement
from app.models.announcement_scope import AnnouncementScopeBinding
from app.models.ai_audit_report import AIAuditReport
from app.models.application import Application
from app.models.application_attachment import ApplicationAttachment
from app.models.archive_record import ArchiveRecord
from app.models.file_info import FileInfo
from app.models.student_report_cache import StudentReportCache
from app.models.user import User
from app.schemas.announcement import AnnouncementCreateRequest, AnnouncementUpdateRequest
from app.services.class_service import get_class_grade, get_class_ids_by_grade, is_graduating_class
from app.services.evaluation_service import build_report_evaluation, build_report_story_copy
from app.services.errors import ServiceError
from app.services.export_workbook_utils import EXPORT_SCORE_COLUMNS, autosize_workbook_columns, build_score_export_columns
from app.services.application_service import get_application_attachments
from app.services.score_summary_service import get_student_score_summary, get_student_score_summary_map, serialize_score_summary
from app.services.serializers import serialize_ai_audit, serialize_announcement, serialize_application, serialize_user
from app.services.system_log_service import write_system_log

CATEGORY_COLORS = {
    "physical_mental": "#c0392b",
    "art": "#7c3aed",
    "labor": "#0f9f7a",
    "innovation": "#2563eb",
}


def create_announcement(db: Session, user: User, payload: AnnouncementCreateRequest) -> dict:
    _require_manage(user)
    archive = _get_archive(db, payload.archive_id)
    scope_rows = _normalize_scope_payload(db, payload)
    announcement = Announcement(
        archive_record_id=archive.id,
        title=payload.title,
        scope_json=json_dumps(_scope_summary(scope_rows)),
        show_fields_json=json_dumps(payload.show_fields or []),
        start_at=payload.start_at,
        end_at=payload.end_at,
        created_by=user.id,
        updated_at=utcnow(),
    )
    db.add(announcement)
    db.flush()
    _replace_scope_bindings(db, announcement, scope_rows)
    db.commit()
    db.refresh(announcement)
    write_system_log(
        db,
        action="announcement.create",
        actor_id=user.id,
        target_type="announcement",
        target_id=str(announcement.id),
    )
    return _serialize_announcement_with_scopes(db, announcement)


def list_announcements(db: Session, user: User) -> list[dict]:
    rows = db.exec(
        select(Announcement, ArchiveRecord)
        .join(ArchiveRecord, Announcement.archive_record_id == ArchiveRecord.id)
        .order_by(Announcement.created_at.desc())
    ).all()
    if user.role == "student":
        rows = [
            (announcement, archive)
            for announcement, archive in rows
            if can_student_view_announcement(announcement, user, db=db)
        ]
    return [_serialize_announcement_with_scopes(db, announcement, fallback_archive=archive) for announcement, archive in rows]


def get_my_announcement_report(db: Session, user: User, announcement_id: int) -> dict:
    announcement = _get_visible_student_announcement(db, user, announcement_id)
    return _get_or_create_student_report_cache(db, user, announcement)


def generate_my_announcement_story_copy(db: Session, user: User, announcement_id: int) -> dict:
    report = get_my_announcement_report(db, user, announcement_id)
    story_copy = report.get("story_copy")
    if not isinstance(story_copy, dict):
        story_copy = {
            "source": "cache",
            "status": "completed",
            "hero_quote": report.get("hero_quote") or "",
            "ending_text": report.get("ending_text") or "",
            "story_cards": report.get("story_cards") or [],
        }
    return {**story_copy, "evaluation": report.get("evaluation"), "cache": report.get("cache")}


def _get_visible_student_announcement(db: Session, user: User, announcement_id: int) -> Announcement:
    if user.role != "student":
        raise ServiceError("permission denied", 1003)
    announcement = db.get(Announcement, announcement_id)
    if not announcement:
        raise ServiceError("announcement not found", 1002)
    if not can_student_view_announcement(announcement, user, db=db):
        raise ServiceError("permission denied", 1003)
    return announcement


def _get_or_create_student_report_cache(db: Session, user: User, announcement: Announcement) -> dict:
    cached = _find_student_report_cache(db, announcement.id, user.id)
    if cached:
        payload = json_loads(cached.report_json, {})
        if isinstance(payload, dict) and payload:
            return _attach_report_cache_meta(payload, cached, hit=True)

    archive = db.get(ArchiveRecord, announcement.archive_record_id)
    generated_at = utcnow()
    report = _build_my_announcement_report_payload(db, user, announcement, archive=archive, generated_at=generated_at)
    cache = StudentReportCache(
        announcement_id=announcement.id,
        student_id=user.id,
        term=archive.term if archive else settings.default_term,
        status="completed",
        report_json=json_dumps(report),
        generated_at=generated_at,
        updated_at=generated_at,
    )
    try:
        db.add(cache)
        db.commit()
        db.refresh(cache)
    except IntegrityError:
        db.rollback()
        cached = _find_student_report_cache(db, announcement.id, user.id)
        if cached:
            payload = json_loads(cached.report_json, {})
            if isinstance(payload, dict) and payload:
                return _attach_report_cache_meta(payload, cached, hit=True)
        raise
    return _attach_report_cache_meta(report, cache, hit=False)


def _find_student_report_cache(db: Session, announcement_id: int | None, student_id: int | None) -> StudentReportCache | None:
    if announcement_id is None or student_id is None:
        return None
    return db.exec(
        select(StudentReportCache).where(
            StudentReportCache.announcement_id == announcement_id,
            StudentReportCache.student_id == student_id,
            StudentReportCache.status == "completed",
        )
    ).first()


def _build_my_announcement_report_payload(
    db: Session,
    user: User,
    announcement: Announcement,
    *,
    archive: ArchiveRecord | None,
    generated_at,
) -> dict:
    score_summary = serialize_score_summary(get_student_score_summary(db, user.id), student_id=user.id)
    radar = _build_radar_payload(score_summary)
    public_applications = [
        application
        for application, student in _query_public_applications(db, announcement)
        if student.id == user.id
    ]
    story_payload = _build_story_payload(public_applications, radar=radar, archive=archive)
    award_history = _build_award_history(db, user, announcement=announcement)
    student_payload = serialize_user(user)
    with ThreadPoolExecutor(max_workers=2) as executor:
        story_future = executor.submit(
            build_report_story_copy,
            student=student_payload,
            radar=radar,
            story_metrics=story_payload["metrics"],
            story_cards=story_payload["cards"],
            award_history=award_history,
        )
        evaluation_future = executor.submit(build_report_evaluation, radar=radar, allow_llm=True)
        story_copy = story_future.result()
        evaluation = evaluation_future.result()
    story_cards = story_copy.get("story_cards") if isinstance(story_copy, dict) else None
    report = {
        "student": student_payload,
        "announcement": _serialize_announcement_with_scopes(db, announcement, fallback_archive=archive),
        "score_summary": score_summary,
        "radar": radar,
        "constellation_items": _build_constellation_items(public_applications),
        "award_history": award_history,
        "story_metrics": story_payload["metrics"],
        "story_cards": story_cards if isinstance(story_cards, list) and story_cards else story_payload["cards"],
        "story_copy": story_copy,
        "hero_quote": story_copy.get("hero_quote") if isinstance(story_copy, dict) else "",
        "ending_text": story_copy.get("ending_text") if isinstance(story_copy, dict) else "",
        "evaluation": evaluation,
        "generated_at": generated_at.isoformat(),
    }
    return report


def _attach_report_cache_meta(report: dict, cache: StudentReportCache, *, hit: bool) -> dict:
    payload = {**report}
    payload["cache"] = {
        "hit": hit,
        "cache_id": cache.id,
        "generated_at": cache.generated_at.isoformat() if cache.generated_at else payload.get("generated_at"),
        "term": cache.term,
        "status": cache.status,
    }
    return payload


def list_announcement_public_applications(
    db: Session,
    user: User,
    announcement_id: int,
    *,
    keyword: str | None = None,
    page: int = 1,
    size: int = 20,
) -> dict:
    if user.role != "student":
        raise ServiceError("permission denied", 1003)
    announcement = db.get(Announcement, announcement_id)
    if not announcement:
        raise ServiceError("announcement not found", 1002)
    if not can_student_view_announcement(announcement, user, db=db):
        raise ServiceError("permission denied", 1003)

    archive = db.get(ArchiveRecord, announcement.archive_record_id)
    normalized_keyword = (keyword or "").strip().lower()
    rows = []
    for application, student in _query_public_applications(db, announcement):
        item = _serialize_public_application(db, application, student)
        if normalized_keyword and not _public_application_matches_keyword(item, normalized_keyword):
            continue
        rows.append(item)

    safe_page = max(int(page or 1), 1)
    safe_size = min(max(int(size or 20), 1), 100)
    start = (safe_page - 1) * safe_size
    return {
        "announcement": _serialize_announcement_with_scopes(db, announcement, fallback_archive=archive),
        "page": safe_page,
        "size": safe_size,
        "total": len(rows),
        "list": rows[start : start + safe_size],
    }


def get_announcement_public_application_detail(
    db: Session,
    user: User,
    announcement_id: int,
    application_id: int,
) -> dict:
    announcement = _get_visible_student_announcement(db, user, announcement_id)
    application, student = _get_public_application_in_announcement(db, announcement, application_id)
    attachments = get_application_attachments(db, application.id)
    for item in attachments:
        if not item.get("file_id"):
            continue
        public_url = (
            f"/api/v1/announcements/{announcement.id}/applications/{application.id}/files/{item['file_id']}"
        )
        item["public_url"] = public_url
        item["url"] = public_url
    payload = serialize_application(application, attachments=attachments, include_detail=True)
    payload["student"] = {
        "id": student.id,
        "name": student.name,
        "account": student.account,
        "class_id": student.class_id,
    }
    payload["student_id"] = student.id
    payload["student_name"] = student.name
    payload["student_account"] = student.account
    payload["class_id"] = student.class_id
    report = db.exec(select(AIAuditReport).where(AIAuditReport.application_id == application.id)).first()
    payload["ai_report"] = serialize_ai_audit(report) if report else None
    return payload


def get_announcement_public_application_file_path(
    db: Session,
    user: User,
    announcement_id: int,
    application_id: int,
    file_id: str,
) -> tuple[Path, str | None]:
    announcement = _get_visible_student_announcement(db, user, announcement_id)
    application, _student = _get_public_application_in_announcement(db, announcement, application_id)
    row = db.exec(
        select(ApplicationAttachment, FileInfo)
        .join(FileInfo, ApplicationAttachment.file_id == FileInfo.id)
        .where(
            ApplicationAttachment.application_id == application.id,
            ApplicationAttachment.file_id == file_id,
            FileInfo.status != "deleted",
        )
    ).first()
    if not row:
        raise ServiceError("file not found", 1002)
    _attachment, file = row
    file_path = Path(file.storage_path)
    if not file_path.exists() or not file_path.is_file():
        raise ServiceError("file not found", 1002)
    return file_path, file.original_name


def update_announcement(db: Session, user: User, announcement_id: int, payload: AnnouncementUpdateRequest) -> dict:
    _require_manage(user)
    announcement = db.get(Announcement, announcement_id)
    if not announcement:
        raise ServiceError("announcement not found", 1002)
    archive = _get_archive(db, payload.archive_id)
    scope_rows = _normalize_scope_payload(db, payload)
    announcement.archive_record_id = archive.id
    announcement.title = payload.title
    announcement.scope_json = json_dumps(_scope_summary(scope_rows))
    announcement.show_fields_json = json_dumps(payload.show_fields or [])
    announcement.start_at = payload.start_at
    announcement.end_at = payload.end_at
    announcement.updated_at = utcnow()
    db.add(announcement)
    _replace_scope_bindings(db, announcement, scope_rows)
    db.commit()
    db.refresh(announcement)
    write_system_log(
        db,
        action="announcement.update",
        actor_id=user.id,
        target_type="announcement",
        target_id=str(announcement.id),
    )
    return _serialize_announcement_with_scopes(db, announcement, fallback_archive=archive)


def close_announcement(db: Session, user: User, announcement_id: int) -> dict:
    _require_manage(user)
    announcement = db.get(Announcement, announcement_id)
    if not announcement:
        raise ServiceError("announcement not found", 1002)
    announcement.status = ANNOUNCEMENT_STATUS_CLOSED
    announcement.closed_at = utcnow()
    announcement.updated_at = utcnow()
    db.add(announcement)
    db.commit()
    archive = db.get(ArchiveRecord, announcement.archive_record_id)
    return _serialize_announcement_with_scopes(db, announcement, fallback_archive=archive)


def reopen_announcement(db: Session, user: User, announcement_id: int) -> dict:
    _require_manage(user)
    announcement = db.get(Announcement, announcement_id)
    if not announcement:
        raise ServiceError("announcement not found", 1002)
    announcement.status = ANNOUNCEMENT_STATUS_ACTIVE
    announcement.closed_at = None
    announcement.updated_at = utcnow()
    db.add(announcement)
    db.commit()
    archive = db.get(ArchiveRecord, announcement.archive_record_id)
    write_system_log(
        db,
        action="announcement.reopen",
        actor_id=user.id,
        target_type="announcement",
        target_id=str(announcement.id),
    )
    return _serialize_announcement_with_scopes(db, announcement, fallback_archive=archive)


def delete_announcement(db: Session, user: User, announcement_id: int) -> None:
    _require_manage(user)
    announcement = db.get(Announcement, announcement_id)
    if not announcement:
        raise ServiceError("announcement not found", 1002)
    db.delete(announcement)
    db.commit()
    write_system_log(
        db,
        action="announcement.delete",
        actor_id=user.id,
        target_type="announcement",
        target_id=str(announcement_id),
    )


def get_announcement_download_path(db: Session, user: User, announcement_id: int) -> Path:
    announcement = db.get(Announcement, announcement_id)
    if not announcement:
        raise ServiceError("announcement not found", 1002)
    if user.role == "student":
        if not can_student_view_announcement(announcement, user, db=db):
            raise ServiceError("permission denied", 1003)
    elif user.role not in {"teacher", "admin"}:
        raise ServiceError("permission denied", 1003)

    settings.export_dir_path.mkdir(parents=True, exist_ok=True)
    if user.role in {"teacher", "admin"}:
        file_path = settings.export_dir_path / f"announcement_{announcement.id}_full.xlsx"
        _build_full_announcement_workbook(db, announcement, file_path)
    else:
        file_path = settings.export_dir_path / f"announcement_{announcement.id}_public.xlsx"
        _build_public_announcement_workbook(db, announcement, file_path)
    return file_path


def _serialize_announcement_with_scopes(
    db: Session,
    announcement: Announcement,
    *,
    fallback_archive: ArchiveRecord | None = None,
) -> dict:
    archive = fallback_archive or db.get(ArchiveRecord, announcement.archive_record_id)
    return serialize_announcement(
        announcement,
        archive.archive_id if archive else "",
        scopes=_serialize_scope_bindings(db, announcement),
    )


def _serialize_scope_bindings(db: Session, announcement: Announcement) -> list[dict]:
    rows = _get_scope_bindings(db, announcement)
    data = []
    for row in rows:
        archive = db.get(ArchiveRecord, row.archive_record_id)
        data.append(
            {
                "id": row.id,
                "archive_id": archive.archive_id if archive else None,
                "grade": row.grade,
                "class_id": row.class_id,
            }
        )
    return data


def _normalize_scope_payload(db: Session, payload: AnnouncementCreateRequest) -> list[dict]:
    raw_scopes = payload.scopes or []
    if not raw_scopes:
        raw_scopes = [
            {
                "archive_id": payload.archive_id,
                "grade": payload.scope.grade,
                "class_ids": payload.scope.class_ids,
            }
        ]

    rows = []
    seen = set()
    for raw in raw_scopes:
        archive_id = raw.archive_id if hasattr(raw, "archive_id") else raw.get("archive_id")
        grade = raw.grade if hasattr(raw, "grade") else raw.get("grade")
        class_ids = raw.class_ids if hasattr(raw, "class_ids") else raw.get("class_ids", [])
        archive = _get_archive(db, archive_id)
        resolved_grade = _resolve_scope_grade(db, grade, class_ids, archive)
        normalized_class_ids = _safe_int_list(class_ids or [])
        if not normalized_class_ids:
            key = (archive.id, resolved_grade, None)
            if key not in seen:
                rows.append({"archive": archive, "grade": resolved_grade, "class_id": None})
                seen.add(key)
            continue
        for class_id in normalized_class_ids:
            class_grade = _resolve_user_grade(db, class_id)
            if class_grade is not None and class_grade != resolved_grade:
                raise ServiceError("class_id does not belong to announcement grade", 1001)
            key = (archive.id, resolved_grade, class_id)
            if key not in seen:
                rows.append({"archive": archive, "grade": resolved_grade, "class_id": class_id})
                seen.add(key)

    grades = {row["grade"] for row in rows}
    if len(grades) != 1:
        raise ServiceError("不同年级公示必须拆分创建", 1001)
    if not rows:
        raise ServiceError("announcement scope is required", 1001)
    return rows


def _replace_scope_bindings(db: Session, announcement: Announcement, scope_rows: list[dict]) -> None:
    existing = db.exec(
        select(AnnouncementScopeBinding).where(AnnouncementScopeBinding.announcement_id == announcement.id)
    ).all()
    for row in existing:
        db.delete(row)

    touched_archives = {announcement.archive_record_id}
    for row in scope_rows:
        archive = row["archive"]
        touched_archives.add(archive.id)
        archive.is_announced = True
        db.add(archive)
        db.add(
            AnnouncementScopeBinding(
                announcement_id=announcement.id,
                archive_record_id=archive.id,
                grade=row["grade"],
                class_id=row["class_id"],
            )
        )

    for archive_id in touched_archives:
        archive = db.get(ArchiveRecord, archive_id)
        if archive:
            archive.is_announced = True
            db.add(archive)


def _scope_summary(scope_rows: list[dict]) -> dict:
    grade = scope_rows[0]["grade"] if scope_rows else None
    class_ids = sorted(row["class_id"] for row in scope_rows if row["class_id"] is not None)
    return {"grade": grade, "class_ids": class_ids}


def _get_scope_bindings(db: Session, announcement: Announcement) -> list[AnnouncementScopeBinding]:
    rows = db.exec(
        select(AnnouncementScopeBinding)
        .where(AnnouncementScopeBinding.announcement_id == announcement.id)
        .order_by(AnnouncementScopeBinding.grade.asc(), AnnouncementScopeBinding.class_id.asc())
    ).all()
    if rows:
        return rows

    archive = db.get(ArchiveRecord, announcement.archive_record_id)
    scope = json_loads(announcement.scope_json, {})
    grade = _resolve_scope_grade(db, scope.get("grade"), scope.get("class_ids") or [], archive)
    class_ids = _safe_int_list(scope.get("class_ids") or [])
    if not class_ids:
        return [
            AnnouncementScopeBinding(
                announcement_id=announcement.id,
                archive_record_id=announcement.archive_record_id,
                grade=grade,
                class_id=None,
            )
        ]
    return [
        AnnouncementScopeBinding(
            announcement_id=announcement.id,
            archive_record_id=announcement.archive_record_id,
            grade=grade,
            class_id=class_id,
        )
        for class_id in class_ids
    ]


def _build_public_announcement_workbook(db: Session, announcement: Announcement, file_path: Path) -> None:
    students = _query_public_students(db, announcement)
    student_ids = [student.id for student in students if student.id is not None]
    score_map = get_student_score_summary_map(db, student_ids)

    wb = Workbook()
    total_ws = wb.active
    total_ws.title = "公示总分"
    total_ws.append(["年级", "班级", "学号", "姓名", "官方总分"])
    for student in students:
        score_summary = serialize_score_summary(score_map.get(student.id), student_id=student.id)
        total_ws.append(
            [
                _resolve_user_grade(db, student.class_id),
                student.class_id,
                student.account,
                student.name,
                score_summary["actual_score"],
            ]
        )

    autosize_workbook_columns(wb)
    wb.save(file_path)


def _build_full_announcement_workbook(db: Session, announcement: Announcement, file_path: Path) -> None:
    students = _query_public_students(db, announcement)
    student_ids = [student.id for student in students if student.id is not None]
    score_map = get_student_score_summary_map(db, student_ids)

    wb = Workbook()
    total_ws = wb.active
    total_ws.title = "完整分数"
    total_ws.append(
        [
            "年级",
            "班级",
            "学号",
            "姓名",
            "官方总分",
            "原始总分",
            *[label for label, _ in EXPORT_SCORE_COLUMNS],
        ]
    )
    for student in students:
        score_summary = serialize_score_summary(score_map.get(student.id), student_id=student.id)
        score_columns = build_score_export_columns(score_summary)
        total_ws.append(
            [
                _resolve_user_grade(db, student.class_id),
                student.class_id,
                student.account,
                student.name,
                score_summary["actual_score"],
                score_summary["raw_total_score"],
                *[score_columns[key] for _, key in EXPORT_SCORE_COLUMNS],
            ]
        )

    application_ws = wb.create_sheet("范围内申报")
    application_ws.append(
        [
            "年级",
            "班级",
            "学号",
            "姓名",
            "申报名称",
            "分类",
            "小类",
            "分数",
            "状态",
            "发生日期",
        ]
    )
    for application, student in _query_public_applications(db, announcement):
        category_rule = SCORE_CATEGORY_RULES.get(application.category, {})
        sub_rule = (category_rule.get("sub_types") or {}).get(application.sub_type, {})
        application_ws.append(
            [
                _resolve_user_grade(db, student.class_id),
                student.class_id,
                student.account,
                student.name,
                application.title,
                category_rule.get("name") or application.category,
                sub_rule.get("name") or application.sub_type,
                application.item_score,
                application.status,
                application.occurred_at.isoformat() if application.occurred_at else "",
            ]
        )
    autosize_workbook_columns(wb)
    wb.save(file_path)


def _query_public_students(db: Session, announcement: Announcement) -> list[User]:
    bindings = _get_scope_bindings(db, announcement)
    class_ids = _expand_scope_class_ids(db, bindings)
    stmt = select(User).where(User.role == "student", User.is_deleted.is_(False))
    if class_ids:
        stmt = stmt.where(User.class_id.in_(class_ids))
    rows = db.exec(stmt.order_by(User.class_id.asc(), User.account.asc())).all()
    grades = {row.grade for row in bindings}
    if class_ids:
        return [row for row in rows if not is_graduating_class(db, row.class_id)]
    return [
        row
        for row in rows
        if _resolve_user_grade(db, row.class_id) in grades and not is_graduating_class(db, row.class_id)
    ]


def _query_public_applications(db: Session, announcement: Announcement) -> list[tuple[Application, User]]:
    students = _query_public_students(db, announcement)
    student_ids = [student.id for student in students if student.id is not None]
    if not student_ids:
        return []
    bindings = _get_scope_bindings(db, announcement)
    archive_by_id = {row.archive_record_id: db.get(ArchiveRecord, row.archive_record_id) for row in bindings}
    rows = db.exec(
        select(Application, User)
        .join(User, Application.applicant_id == User.id)
        .where(
            Application.applicant_id.in_(student_ids),
            Application.is_deleted.is_(False),
            Application.status.in_(("approved", "archived")),
            Application.actual_score_recorded.is_(True),
        )
        .order_by(User.class_id.asc(), User.account.asc(), Application.occurred_at.asc(), Application.id.asc())
    ).all()
    return [
        (application, student)
        for application, student in rows
        if _application_matches_announcement_scope(db, application, student, bindings, archive_by_id)
    ]


def _get_public_application_in_announcement(
    db: Session,
    announcement: Announcement,
    application_id: int,
) -> tuple[Application, User]:
    for application, student in _query_public_applications(db, announcement):
        if application.id == application_id:
            return application, student
    raise ServiceError("该申报不在当前公示范围内", 1003)


def _serialize_public_application(db: Session, application: Application, student: User) -> dict:
    award_rule = serialize_award_rule(application.award_uid)
    category_rule = SCORE_CATEGORY_RULES.get(application.category, {})
    sub_rule = (category_rule.get("sub_types") or {}).get(application.sub_type, {})
    status_label = {
        "approved": "已通过",
        "archived": "已归档",
    }.get(application.status, application.status)
    rule_name = (award_rule or {}).get("rule_name") or "未匹配规则"
    rule_path = (award_rule or {}).get("rule_path") or rule_name
    return {
        "application_id": application.id,
        "student_id": student.id,
        "student_name": student.name,
        "student_account": student.account,
        "class_id": student.class_id,
        "grade": _resolve_user_grade(db, student.class_id),
        "title": application.title,
        "category": application.category,
        "category_name": category_rule.get("name") or application.category,
        "sub_type": application.sub_type,
        "sub_type_name": sub_rule.get("name") or application.sub_type,
        "award_uid": application.award_uid,
        "award_rule": award_rule,
        "award_rule_name": rule_name,
        "award_rule_path": rule_path,
        "score": application.item_score,
        "item_score": application.item_score,
        "status": application.status,
        "status_label": status_label,
        "occurred_at": application.occurred_at.isoformat() if application.occurred_at else None,
        "created_at": application.created_at.isoformat() if application.created_at else None,
    }


def _public_application_matches_keyword(item: dict, keyword: str) -> bool:
    haystacks = [
        item.get("application_id"),
        item.get("student_name"),
        item.get("student_account"),
        item.get("class_id"),
        item.get("grade"),
        item.get("title"),
        item.get("category_name"),
        item.get("sub_type_name"),
        item.get("award_rule_name"),
        item.get("award_rule_path"),
    ]
    return any(keyword in str(value).lower() for value in haystacks if value is not None)


def _expand_scope_class_ids(db: Session, bindings: list[AnnouncementScopeBinding]) -> list[int]:
    result = []
    grades_with_all_classes = {row.grade for row in bindings if row.class_id is None}
    for grade in grades_with_all_classes:
        for class_id in get_class_ids_by_grade(db, grade, include_graduating=False):
            if class_id not in result:
                result.append(class_id)
    for row in bindings:
        if row.class_id is not None and row.class_id not in result:
            result.append(row.class_id)
    return sorted(result)


def _resolve_scope_grade(db: Session, value, class_ids, archive: ArchiveRecord | None) -> int:
    grade = _safe_int(value)
    if grade is not None:
        return grade
    class_id_list = _safe_int_list(class_ids or [])
    class_grades = {_resolve_user_grade(db, class_id) for class_id in class_id_list}
    class_grades.discard(None)
    if len(class_grades) == 1:
        return class_grades.pop()
    if len(class_grades) > 1:
        raise ServiceError("不同年级公示必须拆分创建", 1001)
    if archive and archive.grade:
        return int(archive.grade)
    raise ServiceError("announcement grade is required", 1001)


def _build_radar_payload(score_summary: dict) -> dict:
    sub_scores = score_summary.get("sub_scores") or {}
    category_scores = score_summary.get("category_scores") or {}
    achievement_overflow_scores = score_summary.get("achievement_overflow_scores") or {}
    categories = []
    indicators = []
    for category in SCORE_CATEGORY_KEYS:
        rule = SCORE_CATEGORY_RULES[category]
        color = CATEGORY_COLORS.get(category, "#9c0c13")
        submodules = []
        for sub_type in SCORE_SUB_TYPE_KEYS:
            sub_rule = rule["sub_types"][sub_type]
            field = f"{category}_{sub_type}"
            score = _round_score(sub_scores.get(field, 0.0))
            overflow = _round_score(achievement_overflow_scores.get(f"{category}_achievement_overflow", 0.0)) if sub_type == "achievement" else 0.0
            item = {
                "key": field,
                "category": category,
                "sub_type": sub_type,
                "name": sub_rule["name"],
                "score": score,
                "max_score": _round_score(sub_rule["max_score"]),
                "overflow_score": overflow,
                "score_with_overflow": _round_score(score + overflow),
                "color": color,
            }
            submodules.append(item)
            indicators.append(
                {
                    "key": field,
                    "name": f"{rule['name']}-{sub_rule['name']}",
                    "max": _round_score(sub_rule["max_score"]),
                    "category": category,
                    "color": color,
                }
            )
        categories.append(
            {
                "key": category,
                "name": rule["name"],
                "score": _round_score(category_scores.get(f"{category}_score", 0.0)),
                "max_score": _round_score(rule["max_score"]),
                "color": color,
                "submodules": submodules,
            }
        )
    return {
        "categories": categories,
        "indicators": indicators,
    }


def _build_award_history(db: Session, user: User, *, announcement: Announcement | None = None) -> list[dict]:
    rows = db.exec(
        select(Application).where(
            Application.applicant_id == user.id,
            Application.status.in_(("approved", "archived")),
            Application.actual_score_recorded.is_(True),
            Application.is_deleted.is_(False),
        ).order_by(Application.occurred_at.asc(), Application.created_at.asc(), Application.id.asc())
    ).all()
    history = []
    public_application_ids = None
    if announcement is not None:
        public_application_ids = {
            application.id
            for application, _ in _query_public_applications(db, announcement)
            if application.applicant_id == user.id
        }
    for application in rows:
        if public_application_ids is not None and application.id not in public_application_ids:
            continue
        award_rule = serialize_award_rule(application.award_uid)
        if _is_participation_award(application, award_rule):
            continue
        if _is_hidden_basic_history(application):
            continue
        category_rule = SCORE_CATEGORY_RULES.get(application.category, {})
        sub_rule = (category_rule.get("sub_types") or {}).get(application.sub_type, {})
        history.append(
            {
                "application_id": application.id,
                "title": application.title,
                "occurred_at": application.occurred_at.isoformat() if application.occurred_at else None,
                "status": application.status,
                "score": _round_score(application.item_score or 0.0),
                "category": application.category,
                "category_name": category_rule.get("name") or application.category,
                "sub_type": application.sub_type,
                "sub_type_name": sub_rule.get("name") or application.sub_type,
                "award_uid": application.award_uid,
                "award_rule": award_rule,
                "award_rule_name": (award_rule or {}).get("rule_name"),
            }
        )
    return history


def _build_constellation_items(applications: list[Application]) -> list[dict]:
    result = []
    for application in sorted(applications, key=lambda item: (item.occurred_at, item.created_at, item.id or 0)):
        category_rule = SCORE_CATEGORY_RULES.get(application.category, {})
        sub_rule = (category_rule.get("sub_types") or {}).get(application.sub_type, {})
        award_rule = serialize_award_rule(application.award_uid)
        result.append(
            {
                "application_id": application.id,
                "title": application.title,
                "occurred_at": application.occurred_at.isoformat() if application.occurred_at else None,
                "status": application.status,
                "score": _round_score(application.item_score or 0.0),
                "category": application.category,
                "category_name": category_rule.get("name") or application.category,
                "sub_type": application.sub_type,
                "sub_type_name": sub_rule.get("name") or application.sub_type,
                "award_uid": application.award_uid,
                "award_rule": award_rule,
                "award_rule_name": (award_rule or {}).get("rule_name") or (award_rule or {}).get("rule_path"),
                "color": CATEGORY_COLORS.get(application.category, "#9c0c13"),
            }
        )
    return result


def _build_story_payload(applications: list[Application], *, radar: dict, archive: ArchiveRecord | None) -> dict:
    category_scores = {item.get("key"): item for item in radar.get("categories") or []}
    category_counts = {key: 0 for key in SCORE_CATEGORY_KEYS}
    sub_type_counts = {sub_type: 0 for sub_type in SCORE_SUB_TYPE_KEYS}
    for application in applications:
        if application.category in category_counts:
            category_counts[application.category] += 1
        if application.sub_type in sub_type_counts:
            sub_type_counts[application.sub_type] += 1

    top_category = _pick_top_category(category_scores)
    growth_category = _pick_growth_category(category_scores)
    physical_apps = [item for item in applications if item.category == "physical_mental"]
    art_apps = [item for item in applications if item.category == "art"]
    innovation_apps = [item for item in applications if item.category == "innovation"]
    achievement_apps = [item for item in applications if item.sub_type == "achievement"]
    favorite_season = _favorite_activity_season(physical_apps)
    first_date, last_date = _activity_date_range(applications)
    total_score = sum(float((item.get("score") or 0.0)) for item in category_scores.values())

    metrics = {
        "term": archive.term if archive else None,
        "term_label": format_term_label(archive.term if archive else None),
        "total_applications": len(applications),
        "achievement_applications": len(achievement_apps),
        "art_event_count": len(art_apps),
        "physical_activity_count": len(physical_apps),
        "innovation_count": len(innovation_apps),
        "favorite_sport_season": favorite_season,
        "category_counts": category_counts,
        "sub_type_counts": sub_type_counts,
        "top_category": top_category,
        "growth_category": growth_category,
        "first_activity_date": first_date.isoformat() if first_date else None,
        "last_activity_date": last_date.isoformat() if last_date else None,
    }
    cards = [
        {
            "key": "journey",
            "theme": "overview",
            "eyebrow": metrics["term_label"] or "本次公示",
            "title": "这一学期，你留下了自己的综测足迹",
            "value": len(applications),
            "unit": "条记录",
            "description": _journey_description(applications),
            "color": "#b91c1c",
        },
        {
            "key": "spotlight",
            "theme": "score",
            "eyebrow": "最闪光的方向",
            "title": top_category["name"],
            "value": _round_score(top_category["score"]),
            "unit": "分",
            "description": f"{top_category['name']}是你当前最亮的一束光，已经拿到{_score_text(top_category['score'])}/{_score_text(top_category['max_score'])}分。",
            "color": top_category["color"],
        },
        {
            "key": "art",
            "theme": "art",
            "eyebrow": "文艺片段",
            "title": _art_story_title(len(art_apps)),
            "value": len(art_apps),
            "unit": "次",
            "description": _art_story_description(len(art_apps)),
            "color": CATEGORY_COLORS["art"],
        },
        {
            "key": "sport",
            "theme": "physical",
            "eyebrow": "运动季节",
            "title": _sport_story_title(favorite_season),
            "value": len(physical_apps),
            "unit": "次",
            "description": _sport_story_description(favorite_season, len(physical_apps)),
            "color": CATEGORY_COLORS["physical_mental"],
        },
        {
            "key": "achievement",
            "theme": "achievement",
            "eyebrow": "成果星光",
            "title": "成果与突破把经历连成星轨",
            "value": len(achievement_apps),
            "unit": "项",
            "description": _achievement_story_description(len(achievement_apps), len(innovation_apps)),
            "color": CATEGORY_COLORS["innovation"],
        },
        {
            "key": "growth",
            "theme": "growth",
            "eyebrow": "新的学期",
            "title": f"下一站，试着点亮{growth_category['name']}",
            "value": _round_score(total_score),
            "unit": "总分",
            "description": f"{growth_category['name']}还有继续舒展的空间。新的学期，可以从一个小目标开始，把下一张报告写得更漂亮。",
            "color": growth_category["color"],
        },
    ]
    return {"metrics": metrics, "cards": cards}


def _pick_top_category(category_scores: dict[str, dict]) -> dict:
    fallback = _category_story_item("physical_mental")
    if not category_scores:
        return fallback
    return max(
        (_category_story_item(key, value) for key, value in category_scores.items()),
        key=lambda item: (float(item["score"] or 0.0), float(item["max_score"] or 0.0)),
        default=fallback,
    )


def _pick_growth_category(category_scores: dict[str, dict]) -> dict:
    fallback = _category_story_item("innovation")
    if not category_scores:
        return fallback
    return min(
        (_category_story_item(key, value) for key, value in category_scores.items()),
        key=lambda item: (float(item["score"] or 0.0) / max(float(item["max_score"] or 1.0), 1.0), -float(item["max_score"] or 0.0)),
        default=fallback,
    )


def _category_story_item(category: str, payload: dict | None = None) -> dict:
    rule = SCORE_CATEGORY_RULES.get(category, {})
    payload = payload or {}
    return {
        "key": category,
        "name": payload.get("name") or rule.get("name") or category,
        "score": _round_score(payload.get("score", 0.0)),
        "max_score": _round_score(payload.get("max_score", rule.get("max_score", 0.0))),
        "color": payload.get("color") or CATEGORY_COLORS.get(category, "#9c0c13"),
    }


def _favorite_activity_season(applications: list[Application]) -> str | None:
    counts: dict[str, int] = {}
    for application in applications:
        if not application.occurred_at:
            continue
        season = _season_name(application.occurred_at.month)
        counts[season] = counts.get(season, 0) + 1
    if not counts:
        return None
    order = {"春天": 0, "夏天": 1, "秋天": 2, "冬天": 3}
    return sorted(counts.items(), key=lambda item: (-item[1], order.get(item[0], 99)))[0][0]


def _season_name(month: int) -> str:
    if month in {3, 4, 5}:
        return "春天"
    if month in {6, 7, 8}:
        return "夏天"
    if month in {9, 10, 11}:
        return "秋天"
    return "冬天"


def _activity_date_range(applications: list[Application]):
    dates = sorted(application.occurred_at for application in applications if application.occurred_at)
    if not dates:
        return None, None
    return dates[0], dates[-1]


def _journey_description(applications: list[Application]) -> str:
    if not applications:
        return "这份报告暂时很安静，新的学期可以从第一条申报开始。"
    first_date, last_date = _activity_date_range(applications)
    if first_date and last_date and first_date != last_date:
        return f"从{first_date.month}月到{last_date.month}月，你把学习、活动和探索写进了同一个学期。"
    return "这些记录像一组坐标，标记了你认真参与校园生活的瞬间。"


def _art_story_title(count: int) -> str:
    if count <= 0:
        return "艺术现场还等你打开下一扇门"
    return f"你走过了{count}场文艺现场"


def _art_story_description(count: int) -> str:
    if count <= 0:
        return "下一次讲座、展览、演出或社团活动，都可能成为报告里新的颜色。"
    return "每一次观看、演出、创作或参与，都在训练你发现美、表达美的能力。"


def _sport_story_title(season: str | None) -> str:
    if not season:
        return "你最喜欢的运动季节，等待被记录"
    return f"你最常在{season}运动"


def _sport_story_description(season: str | None, count: int) -> str:
    if count <= 0:
        return "身体的能量会慢慢点亮生活，新的学期可以给自己安排一次轻松的开始。"
    return f"{season or '这个学期'}里的运动记录，说明你正在把健康感和节奏感带进日常。"


def _achievement_story_description(achievement_count: int, innovation_count: int) -> str:
    if achievement_count <= 0:
        return "成果和突破暂时还在路上，先把一次尝试做好，也是一种很好的开始。"
    if innovation_count > 0:
        return f"其中有{innovation_count}项与创新素养有关，说明你已经把想法推进到更远处。"
    return "这些成果记录让努力有了形状，也让下一次尝试看起来更近。"


def _score_text(value) -> str:
    number = float(value or 0.0)
    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _is_participation_award(application: Application, award_rule: dict | None) -> bool:
    haystacks = [
        application.title,
        str(application.award_uid),
        (award_rule or {}).get("rule_name"),
        (award_rule or {}).get("rule_path"),
    ]
    return any("参与未获奖" in str(item) for item in haystacks if item)


def _is_hidden_basic_history(application: Application) -> bool:
    return application.sub_type == "basic" and application.category in {"physical_mental", "art", "labor"}


def _round_score(value) -> float:
    return round(float(value or 0.0), 4)


def can_student_view_announcement(announcement: Announcement, user: User, db: Session | None = None) -> bool:
    if user.role != "student":
        return False
    if announcement.status != ANNOUNCEMENT_STATUS_ACTIVE:
        return False
    now = utcnow()
    start_at = ensure_utc_datetime(announcement.start_at)
    end_at = ensure_utc_datetime(announcement.end_at)
    if start_at and now < start_at:
        return False
    if end_at and now > end_at:
        return False

    if db is None:
        from app.core.database import get_engine

        with Session(get_engine()) as session:
            bindings = _get_scope_bindings(session, announcement)
            if not bindings:
                return False
            if is_graduating_class(session, user.class_id):
                return False
            user_grade = _resolve_user_grade(session, user.class_id)
            for row in bindings:
                if row.class_id is not None and row.class_id == user.class_id:
                    return True
                if row.class_id is None and user_grade == row.grade:
                    return True
            return False
    else:
        bindings = _get_scope_bindings(db, announcement)
    if not bindings:
        return False
    if is_graduating_class(db, user.class_id):
        return False
    user_grade = _resolve_user_grade(db, user.class_id)
    for row in bindings:
        if row.class_id is not None and row.class_id == user.class_id:
            return True
        if row.class_id is None and user_grade == row.grade:
            return True
    return False


def _require_manage(user: User) -> None:
    if user.role not in {"teacher", "admin"}:
        raise ServiceError("permission denied", 1003)


def _get_archive(db: Session, archive_id: str) -> ArchiveRecord:
    archive = db.exec(select(ArchiveRecord).where(ArchiveRecord.archive_id == archive_id)).first()
    if not archive:
        raise ServiceError("archive not found", 1002)
    return archive


def _safe_int(value) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _safe_int_list(values: list) -> list[int]:
    result = []
    for item in values:
        number = _safe_int(item)
        if number is None:
            continue
        if number not in result:
            result.append(number)
    return result


def _application_matches_announcement_scope(
    db: Session,
    application: Application,
    student: User,
    bindings: list[AnnouncementScopeBinding],
    archive_by_id: dict[int, ArchiveRecord | None],
) -> bool:
    student_grade = _resolve_user_grade(db, student.class_id)
    for row in bindings:
        class_matches = row.class_id == student.class_id if row.class_id is not None else row.grade == student_grade
        if not class_matches:
            continue
        archive = archive_by_id.get(row.archive_record_id)
        if archive and not datetime_in_term(application.created_at, archive.term):
            continue
        return True
    return False


def _resolve_user_grade(db: Session, class_id: int | None) -> int | None:
    return get_class_grade(db, class_id)
